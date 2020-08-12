#import pandas
import openpyxl
from pathlib import Path
#import xlrd

xlsx_file = Path("BioGears.xlsx")
wb = openpyxl.load_workbook(xlsx_file, data_only=True)
#of = open("Nutrition.csv","w")
#print(ws['A1'])
#print(ws.cell(row=1,column=1).value)
#print(wb.sheetnames)
#print(wb.active)

ws = wb['Patients']
of = open("Patients.csv","w")
patientsCSV = ""
for row in range(ws.max_row):
	line = ""
	for column in range(ws.max_column):
		cell = ws.cell(row=row+1,column=column+1)
		line += "" if (cell.value is None) else str(cell.value)
		line += ","
	patientsCSV += line[:-1] # Remove the last comma
	patientsCSV += "\n"
of.write(patientsCSV)
of.close()

ws = wb['Substances']
of = open("Substances.csv","w")
substancesCSV = ""
for row in range(ws.max_row):
	line = ""
	for column in range(ws.max_column):
		cell = ws.cell(row=row+1,column=column+1)
		line += "" if (cell.value is None) else str(cell.value)
		line += ","
	substancesCSV += line[:-1] # Remove the last comma
	substancesCSV += "\n"
of.write(substancesCSV)
of.close()

ws = wb['Compounds']
of = open("Compounds.csv","w")
compoundsCSV = ""
for row in range(ws.max_row):
	line = ""
	for column in range(ws.max_column):
		cell = ws.cell(row=row+1,column=column+1)
		line += "" if (cell.value is None) else str(cell.value)
		line += ","
	compoundsCSV += line[:-1] # Remove the last comma
	compoundsCSV += "\n"
of.write(compoundsCSV)
of.close()

ws = wb['Environment']
of = open("Environment.csv","w")
environmentCSV = ""
for row in range(ws.max_row):
	line = ""
	for column in range(ws.max_column):
		cell = ws.cell(row=row+1,column=column+1)
		line += "" if (cell.value is None) else str(cell.value)
		line += ","
	environmentCSV += line[:-1] # Remove the last comma
	environmentCSV += "\n"
of.write(environmentCSV)
of.close()

ws = wb['Nutrition']
of = open("Nutrition.csv","w")
nutritionCSV = ""
for row in range(ws.max_row):
	line = ""
	for column in range(ws.max_column):
		cell = ws.cell(row=row+1,column=column+1)
		line += "" if (cell.value is None) else str(cell.value)
		line += ","
	nutritionCSV += line[:-1] # Remove the last comma
	nutritionCSV += "\n"
of.write(nutritionCSV)
of.close()

ws = wb['Stabilization']
of = open("Stabilization.csv","w")
stabilizationCSV = ""
for row in range(ws.max_row):
	line = ""
	for column in range(ws.max_column):
		cell = ws.cell(row=row+1,column=column+1)
		line += "" if (cell.value is None) else str(cell.value)
		line += ","
	stabilizationCSV += line[:-1] # Remove the last comma
	stabilizationCSV += "\n"
of.write(stabilizationCSV)
of.close()

ws = wb['Tissue']
of = open("Tissue.csv","w")
tissueCSV = ""
for row in range(ws.max_row):
	line = ""
	for column in range(ws.max_column):
		cell = ws.cell(row=row+1,column=column+1)
		line += "" if (cell.value is None) else str(cell.value)
		line += ","
	tissueCSV += line[:-1] # Remove the last comma
	tissueCSV += "\n"
of.write(tissueCSV)
of.close()